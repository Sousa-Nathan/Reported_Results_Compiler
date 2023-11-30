"""
Compiles all the tabs within the Reported Results
workbook into one workbook for greater ease in
data filtering.
"""

import pandas as pd
import PySimpleGUI as sg
import os
import logging
import openpyxl as op
from openpyxl.styles import (Alignment, Border, Font, NamedStyle, PatternFill, Side)

def reported_excel(filepath):
    """
    Formats the workbook for readability.

    Args:
        filepath (str): String filepath to the raw workbook
    """
    
    workbook = op.load_workbook(filename = filepath)
    
    get_sheet = workbook.sheetnames
    
    header = NamedStyle(name = "header")
    header.font = Font(name = "Arial", sz = 8, bold = True)
    header.border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    header.alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    header.fill = PatternFill(fill_type = "solid", start_color = "00538DD5")
    
    formatted_cells = NamedStyle(name = "formatted_cells")
    formatted_cells.font = Font(name = "Arial", sz = 8)
    formatted_cells.border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    formatted_cells.alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    
    for sheet in get_sheet:
        active_sheet = workbook[sheet]
        dimensions = active_sheet.dimensions
        header_row = active_sheet[1]
        
        active_sheet.column_dimensions["A"].width = 14.72
        active_sheet.column_dimensions["B"].width = 12.72
        active_sheet.column_dimensions["C"].width = 17.47
        active_sheet.column_dimensions["D"].width = 12.86
        active_sheet.column_dimensions["E"].width = 10.29
        active_sheet.column_dimensions["F"].width = 9.58
        active_sheet.column_dimensions["G"].width = 9.58
        active_sheet.column_dimensions["H"].width = 8.72
        active_sheet.column_dimensions["I"].width = 8.72
        active_sheet.column_dimensions["J"].width = 8.72
        active_sheet.column_dimensions["K"].width = 9.15
        active_sheet.column_dimensions["L"].width = 10.29
        active_sheet.column_dimensions["M"].width = 10.29
        active_sheet.column_dimensions["N"].width = 10.29
        active_sheet.column_dimensions["O"].width = 10.72
        active_sheet.column_dimensions["P"].width = 10.72
        active_sheet.column_dimensions["Q"].width = 10.72
        active_sheet.column_dimensions["R"].width = 10.72
        active_sheet.column_dimensions["S"].width = 10.72
        active_sheet.column_dimensions["T"].width = 10.72
        active_sheet.column_dimensions["U"].width = 10.72
        
        for row in active_sheet[f"{dimensions}"]:
            for cell in row:
                cell.style = formatted_cells
                cell.style = formatted_cells
        
        wifi_list = ['Wi-Fi 2.4 GHz', 'Wi-Fi 5.2 GHz', 'Wi-Fi 5.3 GHz', 'Wi-Fi 5.5 GHz', 'Wi-Fi 5.8 GHz', 'U-NII 5', 'U-NII 6', 'U-NII 7', 'U-NII 8']
        if any(sheet == active_sheet for sheet in wifi_list):
            for row in active_sheet["L2:M10000"]:
                for cell in row:
                    cell.number_format = "0.00"
        else:
            for row in active_sheet["L2:M10000"]:
                for cell in row:
                    cell.number_format = "0.0"
                    
        for row in active_sheet["N2:U10000"]:
            for cell in row:
                cell.number_format = "0.000"
                
        for cell in header_row:
            cell.style = header
    
    new_sheet = workbook.create_sheet("Author")
    
    new_sheet["A1"] = f"Brought to you by: AJ Newcomer"
    new_sheet["A1"].font = Font(name = "Arial", sz = 72)
    
    workbook.save(filename = filepath)

def concat_data(reported, compiled):
    """
    Compiles the data and generates a formatted workbook
    
    Args:
        reported (str): Directory to the Reported Results workbook
        compiled (str): Directory to the compiled workbook
    """
    
    aj_being_mean = pd.ExcelFile(reported)
    
    sar_sheets = aj_being_mean.sheet_names
    
    remove_list = ["Section 1 Summary", "NFC", "Author"]
    
    sar_sheets = [i for n, i in enumerate(sar_sheets) if i not in remove_list[:n + 1]]
    
    reported_data_set = [pd.read_excel(reported, sheet_name = sar_sheets[tech]) for tech in range(len(sar_sheets))]
    
    for missing in range(len(reported_data_set)):
        reported_data_set[missing].insert(0, "Technology", sar_sheets[missing], True)
    
    smarttx_table = pd.concat(reported_data_set, ignore_index = True)
    smarttx_table = smarttx_table[smarttx_table.Antenna != "Antenna"]
    
    smarttx_table_filtered = smarttx_table[["Technology", "Antenna(s)", "RF Exposure Condition", "Mode(s)", "Power Mode(s)", "Dist. (mm)", "Test Position(s)", "Channel", "Freq. (MHz)", "RB Allocation", "RB Offset", "Max Output Pwr (dBm)", "Meas. (dBm)", "1-g Meas. (W/kg)", "1-g Scaled (W/kg)", "8-g Meas. (W/kg)", "8-g Scaled (W/kg)", "10-g Meas. (W/kg)", "10-g Scaled (W/kg)", "APD Meas. (W/m2)", "APD Scaled (W/m2)"]]
    
    compiled_filepath = os.path.join(compiled, "AJ_is_Mean.xlsx")
    
    with pd.ExcelWriter(f"{compiled_filepath}") as writer:
        smarttx_table_filtered.to_excel(writer, sheet_name = "AJ is Mean", index = False)
    
    reported_excel(compiled_filepath)

def main_window():
    try:
        cwd = os.getcwd()
        
        sg.theme("HotDogStand")
        
        directories = [
            [
                sg.Text("Reported Results Workbook:"),
                sg.Input(key = "-Reported_Results-"),
                sg.FileBrowse(key = "-RR_IN_BROWSE-", file_types = (("Excel Files", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Compiled Workbook Directory:"),
                sg.Input(key = "-Compiled_Directory-"),
                sg.FolderBrowse(key = "-CD_IN_BROWSE-")
            ]
        ]
        
        main_layout = [
            [
                sg.Column(directories, element_justification = "right"),
            ],
            
            [
                sg.Submit(button_text = "Go"),
                sg.Exit()
            ]
        ]
        
        window = sg.Window("AJ is Mean", main_layout)
        
        while True:
            event, values = window.read()
            
            if event in (sg.WIN_CLOSED, "Exit"):
                break
            
            elif event == "Go":
                concat_data(values["-Reported_Results-"], values["-Compiled_Directory-"])
                
                sg.popup(f"Done!\nWorkbook can be found here\n{values['-Compiled_Directory-']}")
    
    except Exception as e:
        cwd_e = os.getcwd()
        
        logger = logging.getLogger("AJ is Mean")
        logger.setLevel(logging.ERROR)
        
        lp = os.path.join(cwd_e, "error.log")
        
        lfh = logging.FileHandler(lp)
        lfh.setLevel(logging.ERROR)
        
        formatter = logging.Formatter("\n%(asctime)s - %(name)s - %(levelname)s - %(message)s")
        lfh.setFormatter(formatter)
        
        logger.addHandler(lfh)
        
        logger.exception(e)
        
        sg.popup_error_with_traceback(f"Error log can be found here: {lp}\n", e)

if __name__ == "__main__":
    main_window()
